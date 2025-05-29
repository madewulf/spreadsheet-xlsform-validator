"""
Validation module for XLSForm and spreadsheet data.
"""
import pandas as pd
import os
import openpyxl
from typing import Dict, List, Any, Tuple, Optional
import re
from elementpath import XPath1Parser, XPathContext
from xml.etree.ElementTree import Element

ERROR_TYPE_MISMATCH = "type_mismatch"
ERROR_CONSTRAINT_UNSATISFIED = "error_constraint_unsatisfied"
ERROR_VALUE_REQUIRED = "error_value_required"

class XLSFormValidator:
    """
    Validator class for XLSForm and spreadsheet data.
    """
    def __init__(self):
        self.survey_sheet = None
        self.choices_sheet = None
        self.question_types = {}
        self.question_constraints = {}
        self.required_questions = set()
        self.choice_lists = {}
        self.question_labels = {}  # Map labels to question names
        self.choice_aliases = {}  # Map list_name to dict of alias -> choice_value

    def parse_xlsform(self, xlsform_file) -> bool:
        """
        Parse the XLSForm file to extract survey and choices sheets.
        
        Args:
            xlsform_file: The XLSForm file object
            
        Returns:
            bool: True if parsing was successful, False otherwise
        """
        try:
            file_path = self._save_temp_file(xlsform_file)
            
            xls = pd.ExcelFile(file_path)
            
            if 'survey' not in xls.sheet_names or 'choices' not in xls.sheet_names:
                return False
                
            self.survey_sheet = pd.read_excel(xls, 'survey')
            
            self.choices_sheet = pd.read_excel(xls, 'choices')
            
            self._process_survey_sheet()
            
            self._process_choices_sheet()
            
            os.remove(file_path)
            
            return True
        except Exception as e:
            print(f"Error parsing XLSForm: {str(e)}")
            return False
    
    def _save_temp_file(self, file_obj) -> str:
        """
        Save a file object to a temporary file.
        
        Args:
            file_obj: The file object
            
        Returns:
            str: The path to the temporary file
        """
        file_path = f"/tmp/{file_obj.name}"
        with open(file_path, 'wb+') as destination:
            for chunk in file_obj.chunks():
                destination.write(chunk)
        return file_path
    
    def _process_survey_sheet(self):
        """
        Process the survey sheet to extract question types and constraints.
        """
        if 'name' not in self.survey_sheet.columns or 'type' not in self.survey_sheet.columns:
            return
        
        for _, row in self.survey_sheet.iterrows():
            if pd.isna(row.get('name')) or pd.isna(row.get('type')):
                continue
                
            name = row['name']
            q_type = row['type']
            
            self.question_types[name] = q_type
            
            if 'label' in self.survey_sheet.columns and not pd.isna(row.get('label')):
                label = row['label']
                self.question_labels[label] = name
            
            if 'required' in self.survey_sheet.columns and row.get('required') == 'yes':
                self.required_questions.add(name)
                
            if 'constraint' in self.survey_sheet.columns and not pd.isna(row.get('constraint')):
                self.question_constraints[name] = row['constraint']
    
    def _process_choices_sheet(self):
        """
        Process the choices sheet to extract choice lists and aliases.
        """
        if 'list_name' not in self.choices_sheet.columns or 'name' not in self.choices_sheet.columns:
            return
            
        for _, row in self.choices_sheet.iterrows():
            if pd.isna(row.get('list_name')) or pd.isna(row.get('name')):
                continue
                
            list_name = row['list_name']
            choice_value = row['name']
            
            if list_name not in self.choice_lists:
                self.choice_lists[list_name] = []
                self.choice_aliases[list_name] = {}
                
            self.choice_lists[list_name].append(choice_value)
            
            if 'alias' in self.choices_sheet.columns and not pd.isna(row.get('alias')):
                alias_value = row['alias']
                self.choice_aliases[list_name][alias_value] = choice_value
    
    def validate_spreadsheet(self, spreadsheet_file, xlsform_data=None) -> Dict[str, Any]:
        """
        Validate a spreadsheet against the XLSForm.
        
        Args:
            spreadsheet_file: The spreadsheet file object
            xlsform_data: Optional XLSForm data if already parsed
            
        Returns:
            Dict: Validation result with 'is_valid' flag and 'errors' list if invalid
        """
        if xlsform_data:
            self.survey_sheet = xlsform_data.get('survey')
            self.choices_sheet = xlsform_data.get('choices')
            self._process_survey_sheet()
            self._process_choices_sheet()
        
        try:
            file_path = self._save_temp_file(spreadsheet_file)
            
            df = pd.read_excel(file_path)
            
            errors = self._validate_spreadsheet_data(df)
            
            os.remove(file_path)
            
            if errors:
                return {
                    'is_valid': False,
                    'errors': errors
                }
            else:
                return {
                    'is_valid': True
                }
        except Exception as e:
            print(f"Error validating spreadsheet: {str(e)}")
            return {
                'is_valid': False,
                'errors': [{
                    'line': 0,
                    'column': 0,
                    'error_type': 'error_parsing',
                    'error_explanation': f"Error parsing spreadsheet: {str(e)}",
                    'question_name': ''
                }]
            }
    
    def _validate_spreadsheet_data(self, df: pd.DataFrame) -> List[Dict[str, Any]]:
        """
        Validate the spreadsheet data against the XLSForm.
        
        Args:
            df: The pandas DataFrame containing the spreadsheet data
            
        Returns:
            List: List of error dictionaries
        """
        errors = []
        
        header_errors = self._validate_headers(df.columns)
        errors.extend(header_errors)
        
        if header_errors:
            return errors
        
        for col_idx, column in enumerate(df.columns):
            question_name = self._resolve_column_to_question_name(column)
            if question_name is None:
                continue
                
            question_type = self.question_types[question_name]
            
            for row_idx, value in enumerate(df[column]):
                if pd.isna(value):
                    if question_name in self.required_questions:
                        errors.append({
                            'line': row_idx + 2,  # +2 because pandas is 0-indexed and Excel has a header row
                            'column': col_idx + 1,
                            'error_type': ERROR_VALUE_REQUIRED,
                            'error_explanation': f"Value is required for question '{question_name}'",
                            'question_name': question_name
                        })
                    continue
                
                type_error = self._validate_type(value, question_type, question_name, list_name=self._extract_list_name(question_type))
                if type_error:
                    errors.append({
                        'line': row_idx + 2,
                        'column': col_idx + 1,
                        'error_type': ERROR_TYPE_MISMATCH,
                        'error_explanation': type_error,
                        'question_name': question_name
                    })
                    continue
                
                if question_name in self.question_constraints:
                    constraint_error = self._validate_constraint(value, self.question_constraints[question_name], question_name)
                    if constraint_error:
                        errors.append({
                            'line': row_idx + 2,
                            'column': col_idx + 1,
                            'error_type': ERROR_CONSTRAINT_UNSATISFIED,
                            'error_explanation': constraint_error,
                            'question_name': question_name
                        })
        
        return errors
    
    def _validate_headers(self, columns) -> List[Dict[str, Any]]:
        """
        Validate that all column headers are present in the XLSForm as names or labels.
        
        Args:
            columns: The column headers from the spreadsheet
            
        Returns:
            List: List of error dictionaries
        """
        errors = []
        
        for col_idx, column in enumerate(columns):
            question_name = self._resolve_column_to_question_name(column)
            if question_name is None:
                errors.append({
                    'line': 1,
                    'column': col_idx + 1,
                    'error_type': ERROR_TYPE_MISMATCH,
                    'error_explanation': f"Column header '{column}' does not match any question name or label in the XLSForm",
                    'question_name': column
                })
        
        return errors
    
    def _validate_type(self, value, question_type: str, question_name: str, list_name: Optional[str] = None) -> Optional[str]:
        """
        Validate a value against a question type.
        
        Args:
            value: The value to validate
            question_type: The question type
            question_name: The question name
            list_name: The list name for select questions
            
        Returns:
            Optional[str]: Error message if validation fails, None otherwise
        """
        if not isinstance(value, str):
            value = str(value)
        
        if question_type == 'integer':
            try:
                int(value)
                return None
            except ValueError:
                return f"Value '{value}' is not a valid integer for question '{question_name}'"
        
        elif question_type == 'decimal':
            try:
                float(value)
                return None
            except ValueError:
                return f"Value '{value}' is not a valid decimal for question '{question_name}'"
        
        elif question_type.startswith('select_one'):
            if list_name and list_name in self.choice_lists:
                value_lower = value.lower()
                choices_lower = [str(choice).lower() for choice in self.choice_lists[list_name]]
                
                aliases_lower = {}
                if list_name in self.choice_aliases:
                    aliases_lower = {str(alias).lower(): choice for alias, choice in self.choice_aliases[list_name].items()}
                
                if value_lower not in choices_lower and value_lower not in aliases_lower:
                    return f"Value '{value}' is not a valid choice for select_one question '{question_name}'"
            return None
        
        elif question_type.startswith('select_multiple'):
            if list_name and list_name in self.choice_lists:
                values = [v.strip().lower() for v in value.split()]
                choices_lower = [str(choice).lower() for choice in self.choice_lists[list_name]]
                
                aliases_lower = set()
                if list_name in self.choice_aliases:
                    aliases_lower = {str(alias).lower() for alias in self.choice_aliases[list_name].keys()}
                
                for v in values:
                    if v not in choices_lower and v not in aliases_lower:
                        return f"Value '{v}' is not a valid choice for select_multiple question '{question_name}'"
            return None
        
        elif question_type == 'date':
            date_pattern = r'^\d{4}-\d{2}-\d{2}$'
            
            if re.match(date_pattern, value):
                return None
                
            try:
                from datetime import datetime
                parsed_date = datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
                date_only = parsed_date.strftime('%Y-%m-%d')
                if re.match(date_pattern, date_only):
                    return None
            except ValueError:
                pass
                
            return f"Value '{value}' is not a valid date (YYYY-MM-DD) for question '{question_name}'"
        
        elif question_type == 'time':
            time_pattern = r'^\d{2}:\d{2}(:\d{2})?$'
            if not re.match(time_pattern, value):
                return f"Value '{value}' is not a valid time (HH:MM[:SS]) for question '{question_name}'"
            return None
        
        return None
    
    def _validate_constraint(self, value, constraint: str, question_name: str) -> Optional[str]:
        """
        Validate a value against a constraint.
        
        Args:
            value: The value to validate
            constraint: The constraint expression
            question_name: The question name
            
        Returns:
            Optional[str]: Error message if validation fails, None otherwise
        """
        
        regex_pattern = r'^regex\(\s*\.\s*,\s*[\'"](.*?)[\'"]\s*\)$'
        regex_match = re.match(regex_pattern, constraint.strip())

        if regex_match:
            pattern_str = regex_match.group(1)
            try:
                str_value = str(value)
                match_result = re.match(pattern_str, str_value)
                print(str_value, match_result)
                if not match_result:
                    return f"Constraint '{constraint}' is not satisfied for value '{value}'"
                return None
            except re.error as e:
                return f"Invalid regex pattern in constraint '{constraint}': {str(e)}"
        
        processed_value = value
        if self.question_types.get(question_name) == 'integer':
            try:
                processed_value = int(value)
            except ValueError:
                return f"Cannot validate constraint for non-integer value '{value}'"
        elif self.question_types.get(question_name) == 'decimal':
            try:
                processed_value = float(value)
            except ValueError:
                return f"Cannot validate constraint for non-decimal value '{value}'"
        
        if self._evaluate_xpath_constraint(constraint, processed_value):
            return None
        else:
            return f"Constraint '{constraint}' is not satisfied for value '{value}'"
    
    def _evaluate_xpath_constraint(self, expression: str, value) -> bool:
        """
        Evaluate an XPath constraint expression.
        
        Args:
            expression: The XPath constraint expression (e.g., ". >= 0 and . < 120")
            value: The value to evaluate against
            
        Returns:
            bool: True if constraint is satisfied, False otherwise
        """
        try:
            fake_node = Element("value")
            fake_node.text = str(value)

            # Create an XPath parser and context
            parser = XPath1Parser()
            tree = parser.parse(expression)
            context = XPathContext(fake_node)

            # Evaluate the constraint expression
            result = tree.evaluate(context)
            return bool(result)
        except Exception:
            return False
            
    def _extract_list_name(self, question_type: str) -> Optional[str]:
        """
        Extract the list name from a select_one or select_multiple question type.
        
        Args:
            question_type: The question type string
            
        Returns:
            Optional[str]: The list name if found, None otherwise
        """
        if question_type.startswith('select_one ') or question_type.startswith('select_multiple '):
            parts = question_type.split(' ', 1)
            if len(parts) > 1:
                return parts[1].strip()
        return None
    def _resolve_column_to_question_name(self, column: str) -> Optional[str]:
        """
        Resolve a column header to a question name, checking both names and labels.
        
        Args:
            column: The column header from the spreadsheet
            
        Returns:
            Optional[str]: The question name if found, None otherwise
        """
        if column in self.question_types:
            return column
        
        if column in self.question_labels:
            return self.question_labels[column]
        
        return None
        
    def create_highlighted_excel(self, spreadsheet_file, errors: List[Dict[str, Any]]) -> str:
        """
        Create an Excel file with highlighted error cells and an error tab.
        
        Args:
            spreadsheet_file: The original spreadsheet file
            errors: List of validation errors
            
        Returns:
            str: Path to the highlighted Excel file
        """
        import tempfile
        from openpyxl.styles import PatternFill
        
        original_path = self._save_temp_file(spreadsheet_file)
        
        wb = openpyxl.load_workbook(original_path)
        ws = wb.active
        
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        
        for error in errors:
            if error['line'] > 1:  # Skip header row
                cell = ws.cell(row=error['line'], column=error['column'])
                cell.fill = red_fill
        
        errors_sheet = wb.create_sheet("Errors")
        errors_sheet.append(["Line", "Column", "Question", "Error Type", "Explanation"])
        
        for error in errors:
            errors_sheet.append([
                error['line'],
                error['column'], 
                error['question_name'],
                error['error_type'],
                error['error_explanation']
            ])
        
        highlighted_path = tempfile.mktemp(suffix='.xlsx')
        wb.save(highlighted_path)
        
        os.remove(original_path)
        
        return highlighted_path
