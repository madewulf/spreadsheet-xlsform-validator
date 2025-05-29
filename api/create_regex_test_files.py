import openpyxl

wb = openpyxl.Workbook()
survey = wb.active
survey.title = "survey"

survey.append(["type", "name", "label", "required", "constraint"])
survey.append(
    [
        "text",
        "month_code",
        "Month Code",
        "yes",
        "regex(.,'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)-\\d{2}$')",
    ]
)
survey.append(
    ["text", "simple_code", "Simple Code", "yes", "regex(.,'^[a-zA-Z0-9]{2}$')"]
)
survey.append(["text", "name", "Name", "yes", ""])

choices = wb.create_sheet("choices")
choices.append(["list_name", "name", "label"])

wb.save("api/test_data/test_xlsform_with_regex.xlsx")

wb = openpyxl.Workbook()
ws = wb.active

ws.append(["month_code", "simple_code", "name"])
ws.append(["Jan-01", "A1", "John Doe"])
ws.append(["Feb-15", "B2", "Jane Smith"])
ws.append(["Dec-31", "Z9", "Alex Johnson"])

wb.save("api/test_data/valid_regex_spreadsheet.xlsx")

wb = openpyxl.Workbook()
ws = wb.active

ws.append(["month_code", "simple_code", "name"])
ws.append(["January-01", "A1", "John Doe"])  # Invalid month format
ws.append(["Feb-1", "ABC", "Jane Smith"])  # Invalid day format and code length
ws.append(["13-31", "1A", "Alex Johnson"])  # Invalid month number

wb.save("api/test_data/invalid_regex_spreadsheet.xlsx")
