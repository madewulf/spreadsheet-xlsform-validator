"""
Tests for the XLSForm validator API.
"""
import os
import tempfile
from django.test import TestCase
from django.urls import reverse
from rest_framework.test import APIClient
from rest_framework import status
import pandas as pd
import openpyxl

class SpreadsheetValidationTests(TestCase):
    """
    Test cases for the spreadsheet validation API.
    """
    def setUp(self):
        """
        Set up test client and create test files.
        """
        self.client = APIClient()
        self.url = reverse('validate-list')
        
        os.makedirs('api/test_data', exist_ok=True)
        
        self.create_test_xlsform()
        
        self.create_valid_test_spreadsheet()
        
        self.create_invalid_test_spreadsheet_type_mismatch()
        
        self.create_invalid_test_spreadsheet_constraint()
        
        self.create_invalid_test_spreadsheet_required()
        
        self.create_valid_test_spreadsheet_with_labels()
        
        self.create_mixed_test_spreadsheet()
    
    def create_test_xlsform(self):
        """
        Create a test XLSForm file with survey and choices sheets.
        """
        wb = openpyxl.Workbook()
        
        survey = wb.active
        survey.title = 'survey'
        
        survey.append(['type', 'name', 'label', 'required', 'constraint'])
        
        survey.append(['integer', 'age', 'Age', 'yes', '. < 150'])
        survey.append(['select_one gender', 'gender', 'Gender', 'yes', ''])
        survey.append(['text', 'name', 'Name', 'yes', ''])
        survey.append(['decimal', 'weight', 'Weight (kg)', 'no', '. > 0'])
        
        choices = wb.create_sheet('choices')
        
        choices.append(['list_name', 'name', 'label'])
        
        choices.append(['gender', 'male', 'Male'])
        choices.append(['gender', 'female', 'Female'])
        choices.append(['gender', 'other', 'Other'])
        
        wb.save('api/test_data/test_xlsform.xlsx')
    
    def create_valid_test_spreadsheet(self):
        """
        Create a valid test spreadsheet that matches the XLSForm.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        
        ws.append(['age', 'gender', 'name', 'weight'])
        
        ws.append([25, 'male', 'John Doe', 75.5])
        ws.append([30, 'female', 'Jane Smith', 65.0])
        ws.append([45, 'other', 'Alex Johnson', 80.2])
        
        wb.save('api/test_data/valid_spreadsheet.xlsx')
    
    def create_invalid_test_spreadsheet_type_mismatch(self):
        """
        Create an invalid test spreadsheet with type mismatches.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        
        ws.append(['age', 'gender', 'name', 'weight'])
        
        ws.append(['twenty-five', 'male', 'John Doe', 75.5])  # age should be integer
        ws.append([30, 'unknown', 'Jane Smith', 65.0])  # gender not in choices
        ws.append([45, 'other', 'Alex Johnson', 'eighty'])  # weight should be decimal
        
        wb.save('api/test_data/invalid_type_mismatch.xlsx')
    
    def create_invalid_test_spreadsheet_constraint(self):
        """
        Create an invalid test spreadsheet with constraint violations.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        
        ws.append(['age', 'gender', 'name', 'weight'])
        
        ws.append([200, 'male', 'John Doe', 75.5])  # age > 150
        ws.append([30, 'female', 'Jane Smith', -5.0])  # weight < 0
        ws.append([45, 'other', 'Alex Johnson', 80.2])
        
        wb.save('api/test_data/invalid_constraint.xlsx')
    
    def create_invalid_test_spreadsheet_required(self):
        """
        Create an invalid test spreadsheet with missing required values.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        
        ws.append(['age', 'gender', 'name', 'weight'])
        
        ws.append([None, 'male', 'John Doe', 75.5])  # missing age
        ws.append([30, None, 'Jane Smith', 65.0])  # missing gender
        ws.append([45, 'other', None, 80.2])  # missing name
        
        wb.save('api/test_data/invalid_required.xlsx')
    
    def test_valid_spreadsheet(self):
        """
        Test validation with a valid spreadsheet.
        """
        with open('api/test_data/test_xlsform.xlsx', 'rb') as xlsform_file:
            with open('api/test_data/valid_spreadsheet.xlsx', 'rb') as spreadsheet_file:
                response = self.client.post(
                    self.url,
                    {
                        'xlsform_file': xlsform_file,
                        'spreadsheet_file': spreadsheet_file
                    },
                    format='multipart'
                )
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['result'], 'valid')
        self.assertNotIn('errors', response.data)
    
    def test_type_mismatch_error(self):
        """
        Test validation with a spreadsheet containing type mismatches.
        """
        with open('api/test_data/test_xlsform.xlsx', 'rb') as xlsform_file:
            with open('api/test_data/invalid_type_mismatch.xlsx', 'rb') as spreadsheet_file:
                response = self.client.post(
                    self.url,
                    {
                        'xlsform_file': xlsform_file,
                        'spreadsheet_file': spreadsheet_file
                    },
                    format='multipart'
                )
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['result'], 'invalid')
        self.assertIn('errors', response.data)
        
        type_mismatch_errors = [e for e in response.data['errors'] if e['error_type'] == 'type_mismatch']
        self.assertTrue(len(type_mismatch_errors) > 0)
    
    def test_constraint_unsatisfied_error(self):
        """
        Test validation with a spreadsheet containing constraint violations.
        """
        with open('api/test_data/test_xlsform.xlsx', 'rb') as xlsform_file:
            with open('api/test_data/invalid_constraint.xlsx', 'rb') as spreadsheet_file:
                response = self.client.post(
                    self.url,
                    {
                        'xlsform_file': xlsform_file,
                        'spreadsheet_file': spreadsheet_file
                    },
                    format='multipart'
                )
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['result'], 'invalid')
        self.assertIn('errors', response.data)
        
        constraint_errors = [e for e in response.data['errors'] if e['error_type'] == 'error_constraint_unsatisfied']
        self.assertTrue(len(constraint_errors) > 0)
    
    def test_value_required_error(self):
        """
        Test validation with a spreadsheet containing missing required values.
        """
        with open('api/test_data/test_xlsform.xlsx', 'rb') as xlsform_file:
            with open('api/test_data/invalid_required.xlsx', 'rb') as spreadsheet_file:
                response = self.client.post(
                    self.url,
                    {
                        'xlsform_file': xlsform_file,
                        'spreadsheet_file': spreadsheet_file
                    },
                    format='multipart'
                )
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['result'], 'invalid')
        self.assertIn('errors', response.data)
        
        required_errors = [e for e in response.data['errors'] if e['error_type'] == 'error_value_required']
        self.assertTrue(len(required_errors) > 0)
    def create_valid_test_spreadsheet_with_labels(self):
        """
        Create a valid test spreadsheet using labels instead of names.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        
        ws.append(['Age', 'Gender', 'Name', 'Weight (kg)'])
        
        ws.append([25, 'male', 'John Doe', 75.5])
        ws.append([30, 'female', 'Jane Smith', 65.0])
        ws.append([45, 'other', 'Alex Johnson', 80.2])
        
        wb.save('api/test_data/valid_spreadsheet_labels.xlsx')
    
    def create_mixed_test_spreadsheet(self):
        """
        Create a test spreadsheet using both names and labels.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        
        ws.append(['age', 'Gender', 'name', 'Weight (kg)'])
        
        ws.append([25, 'male', 'John Doe', 75.5])
        ws.append([30, 'female', 'Jane Smith', 65.0])
        ws.append([45, 'other', 'Alex Johnson', 80.2])
        
        wb.save('api/test_data/mixed_spreadsheet.xlsx')
    
    def test_valid_spreadsheet_with_labels(self):
        """
        Test validation with a spreadsheet using labels as column headers.
        """
        with open('api/test_data/test_xlsform.xlsx', 'rb') as xlsform_file:
            with open('api/test_data/valid_spreadsheet_labels.xlsx', 'rb') as spreadsheet_file:
                response = self.client.post(
                    self.url,
                    {
                        'xlsform_file': xlsform_file,
                        'spreadsheet_file': spreadsheet_file
                    },
                    format='multipart'
                )
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['result'], 'valid')
        self.assertNotIn('errors', response.data)
    
    def test_mixed_names_and_labels(self):
        """
        Test validation with a spreadsheet using both names and labels.
        """
        with open('api/test_data/test_xlsform.xlsx', 'rb') as xlsform_file:
            with open('api/test_data/mixed_spreadsheet.xlsx', 'rb') as spreadsheet_file:
                response = self.client.post(
                    self.url,
                    {
                        'xlsform_file': xlsform_file,
                        'spreadsheet_file': spreadsheet_file
                    },
                    format='multipart'
                )
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['result'], 'valid')
        self.assertNotIn('errors', response.data)
    def test_highlighted_excel_download(self):
        """
        Test downloading highlighted Excel file for invalid spreadsheet.
        """
        with open('api/test_data/test_xlsform.xlsx', 'rb') as xlsform_file:
            with open('api/test_data/invalid_type_mismatch.xlsx', 'rb') as spreadsheet_file:
                response = self.client.post(
                    self.url,
                    {
                        'xlsform_file': xlsform_file,
                        'spreadsheet_file': spreadsheet_file
                    },
                    format='multipart'
                )
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['result'], 'invalid')
        self.assertIn('download_id', response.data)
        
        download_url = reverse('validate-download')
        download_response = self.client.get(
            f"{download_url}?id={response.data['download_id']}"
        )
        
        self.assertEqual(download_response.status_code, 200)
        self.assertEqual(download_response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertIn('highlighted_spreadsheet.xlsx', download_response['Content-Disposition'])
