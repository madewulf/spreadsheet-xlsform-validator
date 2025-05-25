import openpyxl

def create_test_xlsform_with_aliases():
    wb = openpyxl.Workbook()

    survey = wb.active
    survey.title = "survey"

    survey.append(["type", "name", "label", "required", "constraint"])
    survey.append(["integer", "age", "Age", "yes", ". < 150"])
    survey.append(["select_one gender", "gender", "Gender", "yes", ""])
    survey.append(["text", "name", "Name", "yes", ""])

    choices = wb.create_sheet("choices")
    choices.append(["list_name", "name", "label", "alias"])
    choices.append(["gender", "m", "Male", "man"])
    choices.append(["gender", "f", "Female", "woman"])
    choices.append(["gender", "o", "Other", "other_option"])

    wb.save("test_data/test_xlsform_with_aliases.xlsx")

def create_alias_test_spreadsheet():
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(["age", "gender", "name"])
    ws.append([25, "man", "John Doe"])  # using alias "man" instead of "m"
    ws.append([30, "woman", "Jane Smith"])  # using alias "woman" instead of "f" 
    ws.append([45, "other_option", "Alex Johnson"])  # using alias "other_option"

    wb.save("test_data/alias_test_spreadsheet.xlsx")

if __name__ == "__main__":
    create_test_xlsform_with_aliases()
    create_alias_test_spreadsheet()
