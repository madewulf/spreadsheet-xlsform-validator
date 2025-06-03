"""
Tests for the XLSForm validator API.
"""

import os
import tempfile
from django.test import TestCase
from django.urls import reverse
from rest_framework.test import APIClient
from rest_framework import status
from django.core.files.uploadedfile import SimpleUploadedFile
import pandas as pd
import openpyxl
from .validation import XLSFormValidator
import xml.etree.ElementTree as ET
import uuid


class SpreadsheetValidationTests(TestCase):
    """
    Test cases for the spreadsheet validation API.
    """

    def setUp(self):
        """
        Set up test client and create test files.
        """
        self.client = APIClient()
        self.url = reverse("django_xlsform_validator:validate-list")

        os.makedirs("django_xlsform_validator/test_data", exist_ok=True)

        self.create_test_xlsform()
        self.create_test_xlsform_with_integer_choices()

        self.create_valid_test_spreadsheet()

        self.create_invalid_test_spreadsheet_type_mismatch()

        self.create_invalid_test_spreadsheet_constraint()

        self.create_invalid_test_spreadsheet_required()

        self.create_valid_test_spreadsheet_with_labels()

        self.create_mixed_test_spreadsheet()

        self.create_case_insensitive_test_spreadsheet()

        self.create_integer_choice_test_spreadsheet()

        self.create_excel_date_format_spreadsheet()

        self.create_test_xlsform_with_aliases()
        self.create_alias_test_spreadsheet()

        self.create_test_xlsform_with_regex_constraints()
        self.create_valid_regex_test_spreadsheet()
        self.create_invalid_regex_test_spreadsheet()

    def create_test_xlsform(self):
        """
        Create a test XLSForm file with survey and choices sheets.
        """
        wb = openpyxl.Workbook()

        survey = wb.active
        survey.title = "survey"

        survey.append(["type", "name", "label", "required", "constraint"])

        survey.append(["integer", "age", "Age", "yes", ". < 150"])
        survey.append(["select_one gender", "gender", "Gender", "yes", ""])
        survey.append(["text", "name", "Name", "yes", ""])
        survey.append(["decimal", "weight", "Weight (kg)", "no", ". > 0"])

        choices = wb.create_sheet("choices")

        choices.append(["list_name", "name", "label"])

        choices.append(["gender", "male", "Male"])
        choices.append(["gender", "female", "Female"])
        choices.append(["gender", "other", "Other"])

        wb.save("django_xlsform_validator/test_data/test_xlsform.xlsx")

    def create_valid_test_spreadsheet(self):
        """
        Create a valid test spreadsheet that matches the XLSForm.
        """
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.append(["age", "gender", "name", "weight"])

        ws.append([25, "male", "John Doe", 75.5])
        ws.append([30, "female", "Jane Smith", 65.0])
        ws.append([45, "other", "Alex Johnson", 80.2])

        wb.save("django_xlsform_validator/test_data/valid_spreadsheet.xlsx")

    def create_invalid_test_spreadsheet_type_mismatch(self):
        """
        Create an invalid test spreadsheet with type mismatches.
        """
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.append(["age", "gender", "name", "weight"])

        ws.append(["twenty-five", "male", "John Doe", 75.5])  # age should be integer
        ws.append([30, "unknown", "Jane Smith", 65.0])  # gender not in choices
        ws.append([45, "other", "Alex Johnson", "eighty"])  # weight should be decimal

        wb.save("django_xlsform_validator/test_data/invalid_type_mismatch.xlsx")

    def create_invalid_test_spreadsheet_constraint(self):
        """
        Create an invalid test spreadsheet with constraint violations.
        """
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.append(["age", "gender", "name", "weight"])

        ws.append([200, "male", "John Doe", 75.5])  # age > 150
        ws.append([30, "female", "Jane Smith", -5.0])  # weight < 0
        ws.append([45, "other", "Alex Johnson", 80.2])

        wb.save("django_xlsform_validator/test_data/invalid_constraint.xlsx")

    def create_invalid_test_spreadsheet_required(self):
        """
        Create an invalid test spreadsheet with missing required values.
        """
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.append(["age", "gender", "name", "weight"])

        ws.append([None, "male", "John Doe", 75.5])  # missing age
        ws.append([30, None, "Jane Smith", 65.0])  # missing gender
        ws.append([45, "other", None, 80.2])  # missing name

        wb.save("django_xlsform_validator/test_data/invalid_required.xlsx")

    def test_valid_spreadsheet(self):
        """
        Test validation with a valid spreadsheet.
        """
        with open("django_xlsform_validator/test_data/test_xlsform.xlsx", "rb") as xlsform_file:
            with open("django_xlsform_validator/test_data/valid_spreadsheet.xlsx", "rb") as spreadsheet_file:
                response = self.client.post(
                    self.url,
                    {
                        "xlsform_file": xlsform_file,
                        "spreadsheet_file": spreadsheet_file,
                    },
                    format="multipart",
                )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["result"], "valid")
        self.assertNotIn("errors", response.data)

    def test_type_mismatch_error(self):
        """
        Test validation with a spreadsheet containing type mismatches.
        """
        with open("django_xlsform_validator/test_data/test_xlsform.xlsx", "rb") as xlsform_file:
            with open(
                "django_xlsform_validator/test_data/invalid_type_mismatch.xlsx", "rb"
            ) as spreadsheet_file:
                response = self.client.post(
                    self.url,
                    {
                        "xlsform_file": xlsform_file,
                        "spreadsheet_file": spreadsheet_file,
                    },
                    format="multipart",
                )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["result"], "invalid")
        self.assertIn("errors", response.data)

        type_mismatch_errors = [
            e for e in response.data["errors"] if e["error_type"] == "type_mismatch"
        ]
        self.assertTrue(len(type_mismatch_errors) > 0)

    def test_constraint_unsatisfied_error(self):
        """
        Test validation with a spreadsheet containing constraint violations.
        """
        with open("django_xlsform_validator/test_data/test_xlsform.xlsx", "rb") as xlsform_file:
            with open(
                "django_xlsform_validator/test_data/invalid_constraint.xlsx", "rb"
            ) as spreadsheet_file:
                response = self.client.post(
                    self.url,
                    {
                        "xlsform_file": xlsform_file,
                        "spreadsheet_file": spreadsheet_file,
                    },
                    format="multipart",
                )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["result"], "invalid")
        self.assertIn("errors", response.data)

        constraint_errors = [
            e
            for e in response.data["errors"]
            if e["error_type"] == "error_constraint_unsatisfied"
        ]
        self.assertTrue(len(constraint_errors) > 0)

    def test_value_required_error(self):
        """
        Test validation with a spreadsheet containing missing required values.
        """
        with open("django_xlsform_validator/test_data/test_xlsform.xlsx", "rb") as xlsform_file:
            with open("django_xlsform_validator/test_data/invalid_required.xlsx", "rb") as spreadsheet_file:
                response = self.client.post(
                    self.url,
                    {
                        "xlsform_file": xlsform_file,
                        "spreadsheet_file": spreadsheet_file,
                    },
                    format="multipart",
                )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["result"], "invalid")
        self.assertIn("errors", response.data)

        required_errors = [
            e
            for e in response.data["errors"]
            if e["error_type"] == "error_value_required"
        ]
        self.assertTrue(len(required_errors) > 0)

    def create_valid_test_spreadsheet_with_labels(self):
        """
        Create a valid test spreadsheet using labels instead of names.
        """
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.append(["Age", "Gender", "Name", "Weight (kg)"])

        ws.append([25, "male", "John Doe", 75.5])
        ws.append([30, "female", "Jane Smith", 65.0])
        ws.append([45, "other", "Alex Johnson", 80.2])

        wb.save("django_xlsform_validator/test_data/valid_spreadsheet_labels.xlsx")

    def create_mixed_test_spreadsheet(self):
        """
        Create a test spreadsheet using both names and labels.
        """
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.append(["age", "Gender", "name", "Weight (kg)"])

        ws.append([25, "male", "John Doe", 75.5])
        ws.append([30, "female", "Jane Smith", 65.0])
        ws.append([45, "other", "Alex Johnson", 80.2])

        wb.save("django_xlsform_validator/test_data/mixed_spreadsheet.xlsx")

    def test_valid_spreadsheet_with_labels(self):
        """
        Test validation with a spreadsheet using labels as column headers.
        """
        with open("django_xlsform_validator/test_data/test_xlsform.xlsx", "rb") as xlsform_file:
            with open(
                "django_xlsform_validator/test_data/valid_spreadsheet_labels.xlsx", "rb"
            ) as spreadsheet_file:
                response = self.client.post(
                    self.url,
                    {
                        "xlsform_file": xlsform_file,
                        "spreadsheet_file": spreadsheet_file,
                    },
                    format="multipart",
                )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["result"], "valid")
        self.assertNotIn("errors", response.data)

    def test_mixed_names_and_labels(self):
        """
        Test validation with a spreadsheet using both names and labels.
        """
        with open("django_xlsform_validator/test_data/test_xlsform.xlsx", "rb") as xlsform_file:
            with open("django_xlsform_validator/test_data/mixed_spreadsheet.xlsx", "rb") as spreadsheet_file:
                response = self.client.post(
                    self.url,
                    {
                        "xlsform_file": xlsform_file,
                        "spreadsheet_file": spreadsheet_file,
                    },
                    format="multipart",
                )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["result"], "valid")
        self.assertNotIn("errors", response.data)

    def test_highlighted_excel_download_invalid(self):
        """
        Test downloading highlighted Excel file for invalid spreadsheet.
        """
        with open("django_xlsform_validator/test_data/test_xlsform.xlsx", "rb") as xlsform_file:
            with open(
                "django_xlsform_validator/test_data/invalid_type_mismatch.xlsx", "rb"
            ) as spreadsheet_file:
                response = self.client.post(
                    self.url,
                    {
                        "xlsform_file": xlsform_file,
                        "spreadsheet_file": spreadsheet_file,
                    },
                    format="multipart",
                )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["result"], "invalid")
        self.assertIn("download_id", response.data)

        download_url = reverse("django_xlsform_validator:validate-download")
        download_response = self.client.get(
            f"{download_url}?id={response.data['download_id']}"
        )

        self.assertEqual(download_response.status_code, 200)
        self.assertEqual(
            download_response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(
            "highlighted_spreadsheet.xlsx", download_response["Content-Disposition"]
        )

    def create_test_xlsform_with_integer_choices(self):
        """
        Create a test XLSForm file with integer choice values.
        """
        wb = openpyxl.Workbook()

        survey = wb.active
        survey.title = "survey"

        survey.append(["type", "name", "label", "required", "constraint"])

        survey.append(["integer", "age", "Age", "yes", ". < 150"])
        survey.append(["select_one status", "status", "Status", "yes", ""])
        survey.append(["text", "name", "Name", "yes", ""])
        survey.append(["decimal", "weight", "Weight (kg)", "no", ". > 0"])

        choices = wb.create_sheet("choices")

        choices.append(["list_name", "name", "label"])

        choices.append(["status", 1, "Active"])
        choices.append(["status", 2, "Inactive"])
        choices.append(["status", 3, "Pending"])

        wb.save("django_xlsform_validator/test_data/test_xlsform_integer_choices.xlsx")

    def create_integer_choice_test_spreadsheet(self):
        """
        Create a test spreadsheet with integer choice values.
        """
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.append(["age", "status", "name", "weight"])

        ws.append([25, 1, "John Doe", 75.5])  # Integer choice value
        ws.append([30, 2, "Jane Smith", 65.0])  # Integer choice value
        ws.append([45, 3, "Alex Johnson", 80.2])  # Integer choice value

        wb.save("django_xlsform_validator/test_data/integer_choice_spreadsheet.xlsx")

    def create_excel_date_format_spreadsheet(self):
        wb = openpyxl.Workbook()
        survey = wb.active
        survey.title = "survey"

        survey.append(["type", "name", "label", "required", "constraint"])
        survey.append(
            ["date", "last_dispensiation_date", "Last Dispensation Date", "yes", ""]
        )
        survey.append(["text", "name", "Name", "yes", ""])

        choices = wb.create_sheet("choices")
        choices.append(["list_name", "name", "label"])

        wb.save("django_xlsform_validator/test_data/test_xlsform_with_date.xlsx")

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.append(["last_dispensiation_date", "name"])
        ws.append(["2024-09-02 00:00:00", "John Doe"])
        ws.append(["2023-05-15 00:00:00", "Jane Smith"])

        wb.save("django_xlsform_validator/test_data/excel_date_spreadsheet.xlsx")

    def create_case_insensitive_test_spreadsheet(self):
        """
        Create a test spreadsheet with case differences in select_one values.
        """
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.append(["age", "gender", "name", "weight"])

        ws.append([25, "MALE", "John Doe", 75.5])  # MALE instead of male
        ws.append([30, "Female", "Jane Smith", 65.0])  # Female instead of female
        ws.append([45, "oThEr", "Alex Johnson", 80.2])  # oThEr instead of other

        wb.save("django_xlsform_validator/test_data/case_insensitive_spreadsheet.xlsx")

    def test_case_insensitive_validation(self):
        """
        Test that validation works with case-insensitive matching for select_one values.
        """
        with open("django_xlsform_validator/test_data/test_xlsform.xlsx", "rb") as xlsform_file:
            with open(
                "django_xlsform_validator/test_data/case_insensitive_spreadsheet.xlsx", "rb"
            ) as spreadsheet_file:
                response = self.client.post(
                    self.url,
                    {
                        "xlsform_file": xlsform_file,
                        "spreadsheet_file": spreadsheet_file,
                    },
                    format="multipart",
                )

                self.assertEqual(response.status_code, status.HTTP_200_OK)
                self.assertEqual(response.data["result"], "valid")
                self.assertNotIn("errors", response.data)

    def test_highlighted_excel_download(self):
        """
        Test downloading highlighted Excel file for invalid spreadsheet.
        """
        with open("django_xlsform_validator/test_data/test_xlsform.xlsx", "rb") as xlsform_file:
            with open(
                "django_xlsform_validator/test_data/invalid_type_mismatch.xlsx", "rb"
            ) as spreadsheet_file:
                response = self.client.post(
                    self.url,
                    {
                        "xlsform_file": xlsform_file,
                        "spreadsheet_file": spreadsheet_file,
                    },
                    format="multipart",
                )

                self.assertEqual(response.status_code, status.HTTP_200_OK)
                self.assertEqual(response.data["result"], "invalid")
                self.assertIn("download_id", response.data)
                download_url = reverse("django_xlsform_validator:validate-download")
                download_response = self.client.get(
                    f"{download_url}?id={response.data['download_id']}"
                )

                self.assertEqual(download_response.status_code, 200)
                self.assertEqual(
                    download_response["Content-Type"],
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                self.assertIn(
                    "highlighted_spreadsheet.xlsx",
                    download_response["Content-Disposition"],
                )

    def test_integer_choice_validation(self):
        """
        Test that validation works with integer choice values.
        """
        with open(
            "django_xlsform_validator/test_data/test_xlsform_integer_choices.xlsx", "rb"
        ) as xlsform_file:
            with open(
                "django_xlsform_validator/test_data/integer_choice_spreadsheet.xlsx", "rb"
            ) as spreadsheet_file:
                response = self.client.post(
                    self.url,
                    {
                        "xlsform_file": xlsform_file,
                        "spreadsheet_file": spreadsheet_file,
                    },
                    format="multipart",
                )
                self.assertEqual(response.data["result"], "valid")

                self.assertEqual(response.status_code, status.HTTP_200_OK)
                self.assertNotIn("errors", response.data)

    def create_test_xlsform_with_aliases(self):
        """
        Create a test XLSForm file with alias column in choices.
        """
        wb = openpyxl.Workbook()

        survey = wb.active
        survey.title = "survey"

        survey.append(["type", "name", "label", "required", "constraint"])
        survey.append(["integer", "age", "Age", "yes", ". < 150"])
        survey.append(["select_one gender", "gender", "Gender", "yes", ""])
        survey.append(["text", "name", "Name", "yes", ""])

        choices = wb.create_sheet("choices")
        choices.append(["list_name", "name", "label", "alias"])
        choices.append(["gender", "m", "Male", "man"])
        choices.append(["gender", "f", "Female", "woman"])
        choices.append(["gender", "o", "Other", "other_option"])

        wb.save("django_xlsform_validator/test_data/test_xlsform_with_aliases.xlsx")

    def create_alias_test_spreadsheet(self):
        """
        Create a test spreadsheet using alias values.
        """
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.append(["age", "gender", "name"])
        ws.append([25, "man", "John Doe"])  # using alias "man" instead of "m"
        ws.append([30, "woman", "Jane Smith"])  # using alias "woman" instead of "f"
        ws.append([45, "other_option", "Alex Johnson"])  # using alias "other_option"

        wb.save("django_xlsform_validator/test_data/alias_test_spreadsheet.xlsx")

    def test_alias_validation(self):
        """
        Test that validation works with alias values.
        """
        with open("django_xlsform_validator/test_data/test_xlsform_with_aliases.xlsx", "rb") as xlsform_file:
            with open(
                "django_xlsform_validator/test_data/alias_test_spreadsheet.xlsx", "rb"
            ) as spreadsheet_file:
                response = self.client.post(
                    self.url,
                    {
                        "xlsform_file": xlsform_file,
                        "spreadsheet_file": spreadsheet_file,
                    },
                    format="multipart",
                )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["result"], "valid")
        self.assertNotIn("errors", response.data)

    def test_excel_date_format_validation(self):
        """
        Test that validation works with Excel date format (YYYY-MM-DD HH:MM:SS).
        """
        with open("django_xlsform_validator/test_data/test_xlsform_with_date.xlsx", "rb") as xlsform_file:
            with open(
                "django_xlsform_validator/test_data/excel_date_spreadsheet.xlsx", "rb"
            ) as spreadsheet_file:
                response = self.client.post(
                    self.url,
                    {
                        "xlsform_file": xlsform_file,
                        "spreadsheet_file": spreadsheet_file,
                    },
                    format="multipart",
                )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["result"], "valid")
        self.assertNotIn("errors", response.data)

    def create_test_xlsform_with_regex_constraints(self):
        """
        Create a test XLSForm with regex constraints.
        """
        wb = openpyxl.Workbook()
        survey = wb.active
        survey.title = "survey"

        survey.append(["type", "name", "label", "required", "constraint"])
        survey.append(
            [
                "text",
                "month_code",
                "Month Code",
                "yes",
                "regex(.,'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)-\\d{2}$')",
            ]
        )
        survey.append(
            ["text", "simple_code", "Simple Code", "yes", "regex(.,'^[a-zA-Z0-9]{2}$')"]
        )
        survey.append(["text", "name", "Name", "yes", ""])

        choices = wb.create_sheet("choices")
        choices.append(["list_name", "name", "label"])

        wb.save("django_xlsform_validator/test_data/test_xlsform_with_regex.xlsx")

    def create_valid_regex_test_spreadsheet(self):
        """
        Create a valid test spreadsheet with regex constraint values.
        """
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.append(["month_code", "simple_code", "name"])
        ws.append(["Jan-01", "A1", "John Doe"])
        ws.append(["Feb-15", "B2", "Jane Smith"])
        ws.append(["Dec-31", "Z9", "Alex Johnson"])

        wb.save("django_xlsform_validator/test_data/valid_regex_spreadsheet.xlsx")

    def create_invalid_regex_test_spreadsheet(self):
        """
        Create an invalid test spreadsheet with regex constraint violations.
        """
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.append(["month_code", "simple_code", "name"])
        ws.append(["January-01", "A1", "John Doe"])  # Invalid month format
        ws.append(["Feb-1", "ABC", "Jane Smith"])  # Invalid day format and code length
        ws.append(["13-31", "1A", "Alex Johnson"])  # Invalid month number

        wb.save("django_xlsform_validator/test_data/invalid_regex_spreadsheet.xlsx")

    def test_valid_regex_constraint_validation(self):
        """
        Test that validation passes with valid regex constraint values.
        """
        with open("django_xlsform_validator/test_data/test_xlsform_with_regex.xlsx", "rb") as xlsform_file:
            with open(
                "django_xlsform_validator/test_data/valid_regex_spreadsheet.xlsx", "rb"
            ) as spreadsheet_file:
                response = self.client.post(
                    self.url,
                    {
                        "xlsform_file": xlsform_file,
                        "spreadsheet_file": spreadsheet_file,
                    },
                    format="multipart",
                )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["result"], "valid")
        self.assertNotIn("errors", response.data)

    def test_invalid_regex_constraint_validation(self):
        """
        Test that validation fails with invalid regex constraint values.
        """
        with open("django_xlsform_validator/test_data/test_xlsform_with_regex.xlsx", "rb") as xlsform_file:
            with open(
                "django_xlsform_validator/test_data/invalid_regex_spreadsheet.xlsx", "rb"
            ) as spreadsheet_file:
                response = self.client.post(
                    self.url,
                    {
                        "xlsform_file": xlsform_file,
                        "spreadsheet_file": spreadsheet_file,
                    },
                    format="multipart",
                )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["result"], "invalid")
        self.assertIn("errors", response.data)
        self.assertTrue(len(response.data["errors"]) > 0)
        constraint_errors = [
            e
            for e in response.data["errors"]
            if e["error_type"] == "error_constraint_unsatisfied"
        ]
        self.assertTrue(len(constraint_errors) > 0)
        
    def test_regex_validation_with_leading_zeros(self):
        """
        Test that numeric values with leading zeros are correctly validated against regex patterns.
        """
        validator = XLSFormValidator()
        validator.question_types = {'code_ets': 'text'}
        validator.constraints = {'code_ets': "regex(.,'^([0-9]{5})$')"}
        
        # Test that 1652.0 gets formatted as 01652 for 5-digit regex
        error = validator._validate_constraint(1652.0, "regex(.,'^([0-9]{5})$')", 'code_ets')
        self.assertIsNone(error)  # Should pass validation
        
        error = validator._validate_constraint(123.0, "regex(.,'^([0-9]{5})$')", 'code_ets')
        self.assertIsNone(error)  # Should pass validation
        
        error = validator._validate_constraint("01234", "regex(.,'^([0-9]{5})$')", 'code_ets')
        self.assertIsNone(error)  # Should pass validation
        
    def test_regex_constraint_numeric_formatting(self):
        """
        Test specific regex constraint handling for numeric values.
        """
        validator = XLSFormValidator()
        validator.question_types = {'code_ets': 'text'}
        validator.constraints = {'code_ets': "regex(.,'^([0-9]{5})$')"}
        
        error = validator._validate_constraint(1652.0, "regex(.,'^([0-9]{5})$')", 'code_ets')
        self.assertIsNone(error, "Validation should pass for 1652.0 with 5-digit regex pattern")
        
        error = validator._validate_constraint(123.0, "regex(.,'^([0-9]{5})$')", 'code_ets')
        self.assertIsNone(error, "Validation should pass for 123.0 with 5-digit regex pattern")
        
        error = validator._validate_constraint("01234", "regex(.,'^([0-9]{5})$')", 'code_ets')
        self.assertIsNone(error, "Validation should pass for '01234' with 5-digit regex pattern")

    def test_xml_generation_valid_spreadsheet(self):
        """
        Test XML generation with a valid spreadsheet.
        """
        with open("django_xlsform_validator/test_data/test_xlsform.xlsx", "rb") as xlsform_file:
            with open("django_xlsform_validator/test_data/valid_spreadsheet.xlsx", "rb") as spreadsheet_file:
                response = self.client.post(
                    self.url,
                    {
                        "xlsform_file": xlsform_file,
                        "spreadsheet_file": spreadsheet_file,
                        "generate_xml": "true",
                        "version": "2025050304",
                    },
                    format="multipart",
                )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["result"], "valid")
        self.assertIn("xml_files", response.data)
        self.assertIsInstance(response.data["xml_files"], list)
        self.assertEqual(len(response.data["xml_files"]), 3)
        
        for xml_string in response.data["xml_files"]:
            self.assertIn('xmlns:h="http://www.w3.org/1999/xhtml"', xml_string)
            self.assertIn('xmlns:jr="http://openrosa.org/javarosa"', xml_string)
            self.assertIn('version="2025050304"', xml_string)
            self.assertIn('id="file_active_admission"', xml_string)
            self.assertIn("<main_title>", xml_string)
            self.assertIn("<meta>", xml_string)
            self.assertIn("<instanceID>uuid:", xml_string)

    def test_xml_generation_without_flag(self):
        """
        Test that XML is not generated when generate_xml flag is not set.
        """
        with open("django_xlsform_validator/test_data/test_xlsform.xlsx", "rb") as xlsform_file:
            with open("django_xlsform_validator/test_data/valid_spreadsheet.xlsx", "rb") as spreadsheet_file:
                response = self.client.post(
                    self.url,
                    {
                        "xlsform_file": xlsform_file,
                        "spreadsheet_file": spreadsheet_file,
                    },
                    format="multipart",
                )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["result"], "valid")
        self.assertNotIn("xml_files", response.data)

    def test_xml_generation_with_invalid_spreadsheet(self):
        """
        Test that XML generation fails when spreadsheet validation fails.
        """
        with open("django_xlsform_validator/test_data/test_xlsform.xlsx", "rb") as xlsform_file:
            with open("django_xlsform_validator/test_data/invalid_type_mismatch.xlsx", "rb") as spreadsheet_file:
                response = self.client.post(
                    self.url,
                    {
                        "xlsform_file": xlsform_file,
                        "spreadsheet_file": spreadsheet_file,
                        "generate_xml": "true",
                        "version": "2025050304",
                    },
                    format="multipart",
                )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["result"], "invalid")
        self.assertNotIn("xml_files", response.data)
        self.assertIn("errors", response.data)

    def test_xml_generation_default_version(self):
        """
        Test XML generation with default version parameter.
        """
        with open("django_xlsform_validator/test_data/test_xlsform.xlsx", "rb") as xlsform_file:
            with open("django_xlsform_validator/test_data/valid_spreadsheet.xlsx", "rb") as spreadsheet_file:
                response = self.client.post(
                    self.url,
                    {
                        "xlsform_file": xlsform_file,
                        "spreadsheet_file": spreadsheet_file,
                        "generate_xml": "true",
                    },
                    format="multipart",
                )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["result"], "valid")
        self.assertIn("xml_files", response.data)
        
        for xml_string in response.data["xml_files"]:
            self.assertIn('version="1.0"', xml_string)

    def test_xml_generation_with_labels(self):
        """
        Test XML generation with spreadsheet using labels as column headers.
        """
        with open("django_xlsform_validator/test_data/test_xlsform.xlsx", "rb") as xlsform_file:
            with open("django_xlsform_validator/test_data/valid_spreadsheet_labels.xlsx", "rb") as spreadsheet_file:
                response = self.client.post(
                    self.url,
                    {
                        "xlsform_file": xlsform_file,
                        "spreadsheet_file": spreadsheet_file,
                        "generate_xml": "true",
                        "version": "test123",
                    },
                    format="multipart",
                )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["result"], "valid")
        self.assertIn("xml_files", response.data)
        
        for xml_string in response.data["xml_files"]:
            self.assertIn('version="test123"', xml_string)
            self.assertIn("<age>", xml_string)
            self.assertIn("<gender>", xml_string)
            self.assertIn("<name>", xml_string)
            self.assertIn("<weight>", xml_string)

    def test_xml_structure_validation(self):
        """
        Test that generated XML has proper structure and can be parsed.
        """
        with open("django_xlsform_validator/test_data/test_xlsform.xlsx", "rb") as xlsform_file:
            with open("django_xlsform_validator/test_data/valid_spreadsheet.xlsx", "rb") as spreadsheet_file:
                response = self.client.post(
                    self.url,
                    {
                        "xlsform_file": xlsform_file,
                        "spreadsheet_file": spreadsheet_file,
                        "generate_xml": "true",
                        "version": "2025050304",
                    },
                    format="multipart",
                )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        xml_files = response.data["xml_files"]
        
        for xml_string in xml_files:
            try:
                root = ET.fromstring(xml_string)
                
                self.assertEqual(root.tag, "data")
                self.assertEqual(root.get("version"), "2025050304")
                self.assertEqual(root.get("id"), "file_active_admission")
                
                main_title = root.find("main_title")
                self.assertIsNotNone(main_title)
                
                meta = root.find("meta")
                self.assertIsNotNone(meta)
                
                instance_id = meta.find("instanceID")
                self.assertIsNotNone(instance_id)
                self.assertTrue(instance_id.text.startswith("uuid:"))
                
                uuid_part = instance_id.text.replace("uuid:", "")
                uuid.UUID(uuid_part)
                
            except ET.ParseError:
                self.fail(f"Generated XML is not valid: {xml_string}")
            except ValueError:
                self.fail(f"Generated UUID is not valid: {instance_id.text}")

    def test_xml_generation_unit_test(self):
        """
        Test XML generation at the validator level (unit test).
        """
        from django.core.files.uploadedfile import SimpleUploadedFile
        
        validator = XLSFormValidator()
        
        with open("django_xlsform_validator/test_data/test_xlsform.xlsx", "rb") as f:
            xlsform_file = SimpleUploadedFile("test_xlsform.xlsx", f.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.assertTrue(validator.parse_xlsform(xlsform_file))
        
        with open("django_xlsform_validator/test_data/valid_spreadsheet.xlsx", "rb") as f:
            spreadsheet_file = SimpleUploadedFile("valid_spreadsheet.xlsx", f.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            xml_generator = validator.generate_xml_from_spreadsheet(spreadsheet_file, version="unit_test")
            xml_files = list(xml_generator)
            
            self.assertEqual(len(xml_files), 3)
            
            for xml_string in xml_files:
                self.assertIn('version="unit_test"', xml_string)
                self.assertIn("<main_title>", xml_string)
                self.assertIn("<meta>", xml_string)
                
                root = ET.fromstring(xml_string)
                self.assertEqual(root.get("version"), "unit_test")

    def test_xml_generation_iterator_pattern(self):
        """
        Test that XML generation returns an iterator that can be consumed multiple times.
        """
        from django.core.files.uploadedfile import SimpleUploadedFile
        
        validator = XLSFormValidator()
        
        with open("django_xlsform_validator/test_data/test_xlsform.xlsx", "rb") as f:
            xlsform_file = SimpleUploadedFile("test_xlsform.xlsx", f.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.assertTrue(validator.parse_xlsform(xlsform_file))
        
        with open("django_xlsform_validator/test_data/valid_spreadsheet.xlsx", "rb") as f:
            spreadsheet_file = SimpleUploadedFile("valid_spreadsheet.xlsx", f.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            xml_generator = validator.generate_xml_from_spreadsheet(spreadsheet_file, version="iterator_test")
            
            xml_count = 0
            for xml_string in xml_generator:
                xml_count += 1
                self.assertIsInstance(xml_string, str)
                self.assertIn('version="iterator_test"', xml_string)
                
            self.assertEqual(xml_count, 3)

    def test_xml_generation_unique_uuids(self):
        """
        Test that each generated XML has a unique UUID.
        """
        with open("django_xlsform_validator/test_data/test_xlsform.xlsx", "rb") as xlsform_file:
            with open("django_xlsform_validator/test_data/valid_spreadsheet.xlsx", "rb") as spreadsheet_file:
                response = self.client.post(
                    self.url,
                    {
                        "xlsform_file": xlsform_file,
                        "spreadsheet_file": spreadsheet_file,
                        "generate_xml": "true",
                        "version": "uuid_test",
                    },
                    format="multipart",
                )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        xml_files = response.data["xml_files"]
        
        uuids = []
        for xml_string in xml_files:
            root = ET.fromstring(xml_string)
            meta = root.find("meta")
            instance_id = meta.find("instanceID")
            uuid_part = instance_id.text.replace("uuid:", "")
            uuids.append(uuid_part)
        
        self.assertEqual(len(uuids), len(set(uuids)), "All UUIDs should be unique")

    def test_generate_xml_from_dict_basic(self):
        """
        Test basic XML generation from dictionary.
        """
        validator = XLSFormValidator()
        
        with open("django_xlsform_validator/test_data/test_xlsform.xlsx", "rb") as f:
            xlsform_file = SimpleUploadedFile("test_xlsform.xlsx", f.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.assertTrue(validator.parse_xlsform(xlsform_file))
        
        data_dict = {
            "age": 25,
            "gender": "male",
            "name": "John Doe",
            "weight": 70.5
        }
        
        xml_string = validator.generate_xml_from_dict(data_dict, version="dict_test")
        
        self.assertIn('version="dict_test"', xml_string)
        self.assertIn("<main_title>", xml_string)
        self.assertIn("<meta>", xml_string)
        self.assertIn("<age>25</age>", xml_string)
        self.assertIn("<gender>male</gender>", xml_string)
        self.assertIn("<name>John Doe</name>", xml_string)
        self.assertIn("<weight>70.5</weight>", xml_string)
        
        root = ET.fromstring(xml_string)
        self.assertEqual(root.get("version"), "dict_test")
        
        meta = root.find("meta")
        instance_id = meta.find("instanceID")
        self.assertTrue(instance_id.text.startswith("uuid:"))

    def test_generate_xml_from_dict_with_labels(self):
        """
        Test XML generation from dictionary using question labels.
        """
        validator = XLSFormValidator()
        
        with open("django_xlsform_validator/test_data/test_xlsform.xlsx", "rb") as f:
            xlsform_file = SimpleUploadedFile("test_xlsform.xlsx", f.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.assertTrue(validator.parse_xlsform(xlsform_file))
        
        data_dict = {
            "Age": 30,
            "Gender": "female", 
            "Name": "Jane Smith"
        }
        
        xml_string = validator.generate_xml_from_dict(data_dict)
        
        self.assertIn("<age>30</age>", xml_string)
        self.assertIn("<gender>female</gender>", xml_string)
        self.assertIn("<name>Jane Smith</name>", xml_string)

    def test_generate_xml_from_dict_empty_dict(self):
        """
        Test XML generation from empty dictionary.
        """
        validator = XLSFormValidator()
        
        with open("django_xlsform_validator/test_data/test_xlsform.xlsx", "rb") as f:
            xlsform_file = SimpleUploadedFile("test_xlsform.xlsx", f.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.assertTrue(validator.parse_xlsform(xlsform_file))
        
        xml_string = validator.generate_xml_from_dict({})
        
        self.assertIn("<main_title", xml_string)
        self.assertIn("<meta>", xml_string)
        
        root = ET.fromstring(xml_string)
        main_title = root.find("main_title")
        self.assertEqual(len(list(main_title)), 0)

    def test_generate_xml_from_dict_invalid_input(self):
        """
        Test error handling for invalid input types.
        """
        validator = XLSFormValidator()
        
        with self.assertRaises(ValueError) as context:
            validator.generate_xml_from_dict("not a dict")
        
        self.assertIn("data_dict must be a dictionary", str(context.exception))
        
        with self.assertRaises(ValueError) as context:
            validator.generate_xml_from_dict(None)
        
        self.assertIn("data_dict must be a dictionary", str(context.exception))

    def test_generate_xml_from_dict_matches_spreadsheet_output(self):
        """
        Test that dict-based XML generation produces similar output to spreadsheet-based generation.
        """
        validator = XLSFormValidator()
        
        with open("django_xlsform_validator/test_data/test_xlsform.xlsx", "rb") as f:
            xlsform_file = SimpleUploadedFile("test_xlsform.xlsx", f.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.assertTrue(validator.parse_xlsform(xlsform_file))
        
        data_dict = {
            "age": 25,
            "gender": "male",
            "name": "Test User"
        }
        dict_xml = validator.generate_xml_from_dict(data_dict, version="comparison_test")
        
        root = ET.fromstring(dict_xml)
        self.assertEqual(root.tag, "data")
        self.assertEqual(root.get("version"), "comparison_test")
        
        main_title = root.find("main_title")
        self.assertIsNotNone(main_title)
        
        age_elem = main_title.find("age")
        self.assertIsNotNone(age_elem)
        self.assertEqual(age_elem.text, "25")
        
        gender_elem = main_title.find("gender")
        self.assertIsNotNone(gender_elem)
        self.assertEqual(gender_elem.text, "male")
