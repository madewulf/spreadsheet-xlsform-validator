"""
Tests for the XLSForm validator functionality.
"""

import os
import tempfile
from django.test import TestCase
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook

from ..validation import XLSFormValidator


class XLSFormValidatorTests(TestCase):
    """Test the XLSForm validator functionality."""

    def setUp(self):
        """Set up test data."""
        self.validator = XLSFormValidator()
        
        self.xlsform_file = self._create_test_xlsform()
        
        self.valid_spreadsheet = self._create_valid_spreadsheet()
        
        self.invalid_spreadsheet = self._create_invalid_spreadsheet()

    def _create_test_xlsform(self):
        """Create a test XLSForm file."""
        wb = Workbook()
        
        survey = wb.active
        survey.title = "survey"
        survey.append(["type", "name", "label", "required", "constraint"])
        survey.append(["integer", "age", "Age", "yes", ". >= 18 and . <= 100"])
        survey.append(["select_one yes_no", "consent", "Consent", "yes", ""])
        
        choices = wb.create_sheet("choices")
        choices.append(["list_name", "name", "label"])
        choices.append(["yes_no", "yes", "Yes"])
        choices.append(["yes_no", "no", "No"])
        
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        wb.save(path)
        
        with open(path, "rb") as f:
            content = f.read()
        
        os.unlink(path)
        return SimpleUploadedFile("test_xlsform.xlsx", content)

    def _create_valid_spreadsheet(self):
        """Create a valid test spreadsheet."""
        wb = Workbook()
        ws = wb.active
        
        ws.append(["age", "consent"])
        ws.append(["25", "yes"])
        ws.append(["30", "no"])
        
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        wb.save(path)
        
        with open(path, "rb") as f:
            content = f.read()
        
        os.unlink(path)
        return SimpleUploadedFile("valid_data.xlsx", content)

    def _create_invalid_spreadsheet(self):
        """Create an invalid test spreadsheet (type mismatch)."""
        wb = Workbook()
        ws = wb.active
        
        ws.append(["age", "consent"])
        ws.append(["not_a_number", "yes"])
        ws.append(["30", "invalid_choice"])
        
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        wb.save(path)
        
        with open(path, "rb") as f:
            content = f.read()
        
        os.unlink(path)
        return SimpleUploadedFile("invalid_data.xlsx", content)

    def test_parse_xlsform(self):
        """Test parsing an XLSForm file."""
        result = self.validator.parse_xlsform(self.xlsform_file)
        self.assertTrue(result)
        self.assertIn("age", self.validator.questions)
        self.assertIn("consent", self.validator.questions)
        self.assertEqual(self.validator.questions["age"]["type"], "integer")
        self.assertEqual(self.validator.questions["consent"]["type"], "select_one")

    def test_validate_valid_spreadsheet(self):
        """Test validating a valid spreadsheet."""
        self.validator.parse_xlsform(self.xlsform_file)
        result = self.validator.validate_spreadsheet(self.valid_spreadsheet)
        self.assertTrue(result["is_valid"])
        self.assertEqual(len(result["errors"]), 0)

    def test_validate_invalid_spreadsheet(self):
        """Test validating an invalid spreadsheet."""
        self.validator.parse_xlsform(self.xlsform_file)
        result = self.validator.validate_spreadsheet(self.invalid_spreadsheet)
        self.assertFalse(result["is_valid"])
        self.assertGreater(len(result["errors"]), 0)
        
        type_mismatch_errors = [e for e in result["errors"] if e["error_type"] == "type_mismatch"]
        self.assertGreater(len(type_mismatch_errors), 0)
