"""
Validation module for XLSForm and spreadsheet data.
"""

import io
import json
import tempfile
from cmath import isnan
from typing import Dict, List, Any, Optional

import openpyxl
import pandas as pd
from pandas import Timestamp
from pyxform import create_survey_from_xls, errors

from xlsform_validator.odk_validate import check_xform, ODKValidateError


class NamedBytesIO(io.BytesIO):
    """
    BytesIO wrapper that provides a .name attribute for compatibility with libraries
    that expect file-like objects with names (like pyxform).
    """

    def __init__(self, initial_bytes=None, name="temp_file"):
        if initial_bytes is not None:
            super().__init__(initial_bytes)
        else:
            super().__init__()
        self.name = name


ERROR_TYPE_MISMATCH = "type_mismatch"
ERROR_CONSTRAINT_UNSATISFIED = "Answer is violating a constraint"
ERROR_VALUE_REQUIRED = "Answer was required but empty"


class XLSFormValidator:
    """
    Validator class for XLSForm and spreadsheet data.
    """

    def __init__(self):
        self.question_types: dict[str, str] = {}
        self.choice_lists: dict[str, list[str]] = {}
        self.question_labels: dict[str, str] = {}  # Map labels to question names
        self.choice_aliases: dict[str, dict[str, str]] = {}  # Map list_name to dict of alias -> choice_value
        self.question_constraints = {}
        self.question_constraint_messages = {}
        self.survey_xml: Optional[str] = None  # Store the XLSForm XML structure

    def parse_xlsform(self, xlsform_file, version: str = '1.0') -> bool:
        """
        Parse the XLSForm file using pyxform to extract survey structure.

        Args:
            xlsform_file: The XLSForm file object
            version: Version string for the XML files

        Returns:
            bool: True if parsing was successful, False otherwise
        """
        try:
            memory_file = self._save_temp_file(xlsform_file)
            survey = create_survey_from_xls(memory_file)
            survey.version = version
            survey_json = survey.to_json()

            self.survey_xml = survey.to_xml(validate=False)
            parsed_survey = json.loads(survey_json)
            self._extract_questions_from_pyxform(parsed_survey)

            self._extract_choices_from_pyxform(parsed_survey)

            return True
        except errors.PyXFormError as e:
            print(f"Error parsing XLSForm with pyxform: {str(e)}")
            return False
        except Exception as e:
            print(f"Error parsing XLSForm: {str(e)}")
            return False

    @staticmethod
    def _save_temp_file(file_obj) -> NamedBytesIO:
        """
        Save uploaded file content to an in-memory BytesIO object.

        Args:
            file_obj: Django UploadedFile object

        Returns:
            NamedBytesIO: In-memory file object with the file content
        """
        file_obj.seek(0)
        if hasattr(file_obj, "chunks"):
            file_content = b"".join(chunk for chunk in file_obj.chunks())
        else:
            file_content = file_obj.read()
        file_obj.seek(0)

        memory_file = NamedBytesIO(file_content, name=file_obj.name)
        memory_file.seek(0)

        return memory_file

    def _extract_questions_from_pyxform(self, parsed_survey: dict):
        """
        Extract question types, labels, constraints, and required fields from pyxform structure.

        Args:
            parsed_survey: The parsed pyxform JSON structure
        """
        if "children" not in parsed_survey:
            return

        for child in parsed_survey["children"]:
            self._process_question_node(child)

    def _process_question_node(self, node: dict):
        """
        Process a single question node from pyxform structure.

        Args:
            node: A question node from the pyxform children array
        """
        if "name" not in node or "type" not in node:
            return

        name = node["name"]
        q_type = node["type"]
        if q_type.startswith("select"):
            q_type = q_type + " " + node["list_name"]
        if name == "meta" or q_type == "group":
            if "children" in node:
                for child in node["children"]:
                    self._process_question_node(child)
            return

        self.question_types[name] = q_type

        if "label" in node:
            label = node["label"]
            self.question_labels[label] = name

        if "bind" in node and "constraint" in node["bind"]:
            self.question_constraints[name] = node["bind"]["constraint"]
            if "jr:constraintMsg" in node["bind"]:
                self.question_constraint_messages[name] = node["bind"]["jr:constraintMsg"]

    def _extract_choices_from_pyxform(self, parsed_survey: dict):
        """
        Extract choice lists from pyxform structure.

        Args:
            parsed_survey: The parsed pyxform JSON structure
        """
        if "choices" not in parsed_survey:
            return

        for list_name, choices in parsed_survey["choices"].items():
            if list_name not in self.choice_lists:
                self.choice_lists[list_name] = []
                self.choice_aliases[list_name] = {}

            for choice in choices:
                if "name" in choice:
                    choice_value = choice["name"]
                    self.choice_lists[list_name].append(choice_value)

                    if "alias" in choice:
                        alias_value = choice["alias"]
                        self.choice_aliases[list_name][alias_value] = choice_value

    def validate_spreadsheet(
            self, spreadsheet_file
    ) -> Dict:
        """
        Validate a spreadsheet against the XLSForm.

        Args:
            spreadsheet_file: The spreadsheet file object

        Returns:
            Dict: Validation result with 'is_valid' flag and 'errors' list if invalid
        """

        memory_file = self._save_temp_file(spreadsheet_file)
        memory_file.seek(0)

        df = pd.read_excel(memory_file, dtype=str)

        try:
            results = self._validate_spreadsheet_data(df)
        except Exception:
            return {
                "is_valid": False,
                "errors": [
                    {
                        "line": 0,
                        "column": 0,
                        "error_type": "error_parsing",
                        "error_explanation": "Failed to parse XLSForm file. Make sure it contains 'survey' and 'choices' sheets.",
                        "question_name": "",
                    }
                ],
            }

        valides = list(filter(lambda x: 'xml' in x, results))
        errors = list(filter(lambda x: 'xml' not in x, results))

        return {
            "is_valid": len(errors) == 0,
            "errors": errors,
            "valides": valides,
        }

    def _validate_spreadsheet_data(self, df: pd.DataFrame) -> List[Dict]:
        """
        Validate the spreadsheet data against the XLSForm.

        Args:
            df: The pandas DataFrame containing the spreadsheet data

        Returns:
            List: List of error dictionaries
        """
        results = []
        answers = {}
        for col_idx, column in enumerate(df.columns):
            question_name = self._resolve_column_to_question_name(column)
            if question_name is None:
                continue

            for row_idx, value in enumerate(df[column]):
                existing = answers.get(row_idx)
                if not existing:
                    answers[row_idx] = {}
                    existing = answers[row_idx]
                existing[question_name] = self._format_value(question_name, value)

        xml_file = tempfile.NamedTemporaryFile(delete=False)
        with open(xml_file.name, 'w') as f:
            f.write(self.survey_xml)
        for i, answer in enumerate(answers):
            try:
                results.append({
                    'xml': check_xform(xml_file.name, json.dumps(answers[answer])),
                })
            except ODKValidateError as e:
                error_json = json.loads(str(e))
                results.append({
                    'line': i + 2,
                    'column': list(self.question_labels.values()).index(error_json["question"]) + 1,
                    'question_name': error_json["question"],
                    'error_type': error_json["error"],
                    'error_explanation': self._get_error_explanation(error_json),
                    'constraint_message': self.question_constraint_messages.get(error_json["question"]),
                })

        return results

    def _resolve_column_to_question_name(self, column: str) -> Optional[str]:
        """
        Resolve a column header to a question name, checking both names and labels.

        Args:
            column: The column header from the spreadsheet

        Returns:
            Optional[str]: The question name if found, None otherwise
        """
        if column in self.question_types:
            return column

        if column in self.question_labels:
            return self.question_labels[column]

        return None

    def _format_value(self, question_name: str, value: str) -> str:
        """
        format a given value so that it can be converted to a JSON string
        """
        if isinstance(value, Timestamp):
            return value.__format__("%d/%m/%YT%H:%M:%S")

        question_type = self.question_types.get(question_name)
        if question_type.startswith("select one"):
            list_name = self._extract_list_name(question_type)
            if list_name and list_name in self.choice_lists:
                return self._get_choice_from_value(
                    self.choice_lists[list_name],
                    self.choice_aliases.get(list_name),
                    value,
                )
        if question_type.startswith("select multiple"):
            list_name = self._extract_list_name(question_type)
            if list_name and list_name in self.choice_lists:
                choice_list = self.choice_lists[list_name],
                choice_aliases = self.choice_aliases.get(list_name)
                map(lambda x: self._get_choice_from_value(choice_list, choice_aliases, x), value.split(','))

        return str(value) if isinstance(value, str) or not isnan(value) else ""

    @staticmethod
    def _get_choice_from_value(choice_list: list[str], choice_aliases: Optional[dict[str, str]], value: str) -> str:
        value_lower = value.lower() if isinstance(value, str) else value
        choices_lower = [str(choice).lower() for choice in choice_list]
        if value_lower in choices_lower:
            return choice_list[choices_lower.index(value_lower)]

        if choice_aliases:
            aliases_lower = {
                str(alias).lower(): choice
                for alias, choice in choice_aliases.items()
            }

            if value_lower in aliases_lower:
                return aliases_lower[value_lower]

        return value

    def create_highlighted_excel(
            self, spreadsheet_file, errors: List[Dict[str, Any]]
    ) -> io.BytesIO:
        """
        Create an Excel file with highlighted error cells and an error tab.

        Args:
            spreadsheet_file: The original spreadsheet file
            errors: List of validation errors

        Returns:
            io.BytesIO: In-memory Excel file with highlighted errors
        """
        from openpyxl.styles import PatternFill

        memory_file = self._save_temp_file(spreadsheet_file)
        memory_file.seek(0)

        wb = openpyxl.load_workbook(memory_file)
        ws = wb.active

        red_fill = PatternFill(
            start_color="FF0000", end_color="FF0000", fill_type="solid"
        )

        for error in errors:
            if error["line"] > 1:
                cell = ws.cell(row=error["line"], column=error["column"])
                if cell is not None:
                    cell.fill = red_fill

        errors_sheet = wb.create_sheet("Errors")
        errors_sheet.append(["Line", "Column", "Question", "Error Type", "Explanation", "Constraint Message"])

        for error in errors:
            constraint_message = error.get("constraint_message", "")
            errors_sheet.append(
                [
                    error["line"],
                    error["column"],
                    error["question_name"],
                    error["error_type"],
                    error["error_explanation"],
                    constraint_message,
                ]
            )

        output_buffer = io.BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)

        return output_buffer

    @staticmethod
    def _extract_list_name(question_type: str) -> Optional[str]:
        """
        Extract the list name from a select_one or select_multiple question type.

        Args:
            question_type: The question type string

        Returns:
            Optional[str]: The list name if found, None otherwise
        """
        if question_type.startswith("select one ") or question_type.startswith(
                "select multiple "
        ):
            parts = question_type.split(" ", 2)
            if len(parts) > 2:
                return parts[2].strip()
        return None

    def generate_xml_from_spreadsheet(
            self, spreadsheet_file
    ):
        """
        Generate XML files from validated spreadsheet data.

        Args:
            spreadsheet_file: The spreadsheet file object

        Returns:
            Iterator yielding XML strings for each row
        """
        validation_result = self.validate_spreadsheet(spreadsheet_file)
        if not validation_result["is_valid"]:
            raise ValueError(
                f"Spreadsheet validation failed: {validation_result}"
            )

        for result in validation_result["valides"]:
            yield result["xml"]

    def _get_error_explanation(self, error_json: dict[str, str]) -> str:
        if error_json["error"] == ERROR_CONSTRAINT_UNSATISFIED:
            return f"Constraint '{self.question_constraints[error_json['question']]}' is not satisfied for value '{error_json['answer']}'"
        if error_json["error"] == ERROR_VALUE_REQUIRED:
            return f"Value is required for question '{error_json['question']}'"
        return error_json["error"]
