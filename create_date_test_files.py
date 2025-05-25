import openpyxl

wb = openpyxl.Workbook()
survey = wb.active
survey.title = 'survey'

survey.append(['type', 'name', 'label', 'required', 'constraint'])
survey.append(['date', 'last_dispensiation_date', 'Last Dispensation Date', 'yes', ''])
survey.append(['text', 'name', 'Name', 'yes', ''])

choices = wb.create_sheet('choices')
choices.append(['list_name', 'name', 'label'])

wb.save('api/test_data/test_xlsform_with_date.xlsx')

wb = openpyxl.Workbook()
ws = wb.active

ws.append(['last_dispensiation_date', 'name'])
ws.append(['2024-09-02 00:00:00', 'John Doe'])
ws.append(['2023-05-15 00:00:00', 'Jane Smith'])

wb.save('api/test_data/excel_date_spreadsheet.xlsx')
